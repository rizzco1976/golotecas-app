import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import io
import re
from pathlib import Path

# ── Configuración ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Golotecas · Análisis de Ventas",
    page_icon="🍬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Estilos ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.main { background-color: #0f1117; }

.metric-card {
    background: linear-gradient(135deg, #1a1d2e 0%, #16192b 100%);
    border: 1px solid #2a2d3e;
    border-radius: 12px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 0.5rem;
}
.metric-label {
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #6b7280;
    margin-bottom: 0.3rem;
}
.metric-value {
    font-size: 2rem;
    font-weight: 700;
    color: #f9fafb;
    line-height: 1;
}
.metric-delta {
    font-size: 0.78rem;
    margin-top: 0.25rem;
}
.delta-pos { color: #34d399; }
.delta-neg { color: #f87171; }
.delta-neu { color: #9ca3af; }

.section-title {
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #6366f1;
    padding: 0.5rem 0 0.8rem 0;
    border-bottom: 1px solid #1f2937;
    margin-bottom: 1rem;
}

.producto-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 0.55rem 0.8rem;
    border-radius: 8px;
    margin-bottom: 0.3rem;
    background: #1a1d2e;
    border-left: 3px solid #6366f1;
}
.producto-nombre {
    font-size: 0.78rem;
    color: #d1d5db;
    flex: 1;
}
.producto-valor {
    font-size: 0.85rem;
    font-weight: 600;
    color: #a5b4fc;
    margin-left: 1rem;
    white-space: nowrap;
}

.proyeccion-card {
    background: linear-gradient(135deg, #1e1b4b 0%, #1a1d2e 100%);
    border: 1px solid #4338ca;
    border-radius: 12px;
    padding: 1.4rem;
    margin-bottom: 1rem;
}
.proyeccion-title {
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #818cf8;
    margin-bottom: 0.8rem;
}
.proyeccion-valor {
    font-size: 2.4rem;
    font-weight: 700;
    color: #a5b4fc;
}
.proyeccion-rango {
    font-size: 0.78rem;
    color: #6b7280;
    margin-top: 0.2rem;
}

.tag-verde {
    display: inline-block;
    background: rgba(52,211,153,0.15);
    color: #34d399;
    font-size: 0.68rem;
    font-weight: 600;
    padding: 0.15rem 0.55rem;
    border-radius: 20px;
    letter-spacing: 0.05em;
}
.tag-rojo {
    display: inline-block;
    background: rgba(248,113,113,0.15);
    color: #f87171;
    font-size: 0.68rem;
    font-weight: 600;
    padding: 0.15rem 0.55rem;
    border-radius: 20px;
    letter-spacing: 0.05em;
}
.tag-amarillo {
    display: inline-block;
    background: rgba(251,191,36,0.15);
    color: #fbbf24;
    font-size: 0.68rem;
    font-weight: 600;
    padding: 0.15rem 0.55rem;
    border-radius: 20px;
    letter-spacing: 0.05em;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: #0d1018;
    border-right: 1px solid #1f2937;
}

/* Ocultar elementos default de Streamlit */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ── Constantes ─────────────────────────────────────────────────────────────────
MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
         'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

MESES_COL = {
    'Enero': 2, 'Febrero': 3, 'Marzo': 4,
    'Abril': 6, 'Mayo': 7, 'Junio': 8,
    'Julio': 10, 'Agosto': 11, 'Septiembre': 12,
    'Octubre': 14, 'Noviembre': 15, 'Diciembre': 16
}

TRIMESTRE_DE_MES = {
    'Enero': 1, 'Febrero': 1, 'Marzo': 1,
    'Abril': 2, 'Mayo': 2, 'Junio': 2,
    'Julio': 3, 'Agosto': 3, 'Septiembre': 3,
    'Octubre': 4, 'Noviembre': 4, 'Diciembre': 4,
}

# ── Funciones de carga y parseo ────────────────────────────────────────────────
NEGOCIO_PATTERN = re.compile(r'^\d{2} - ')
CATEGORIA_PATTERN = re.compile(r'^\d{4} - ')


@st.cache_data(show_spinner=False)
def cargar_datos(archivo_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    registros = []
    cuenta_actual = None
    negocio_actual = None
    categoria_actual = None

    for row in rows:
        col_b = str(row[1]).strip() if row[1] is not None else ''
        col_a = row[0]

        # Detectar bloque de cuenta
        if col_b.startswith('cuenta'):
            cuenta_actual = col_b
            negocio_actual = None
            categoria_actual = None
            continue

        if not cuenta_actual:
            continue

        # Detectar negocio (2 dígitos) o categoría (4 dígitos)
        if NEGOCIO_PATTERN.match(col_b):
            negocio_actual = col_b
            categoria_actual = None
            continue
        if CATEGORIA_PATTERN.match(col_b):
            categoria_actual = col_b
            continue

        # Filas de producto (col_a es código numérico)
        if col_a and str(col_a).strip() not in ['', 'None', 'Codigo', 'codigo', 'cuenta', '#N/A']:
            try:
                codigo = int(float(str(col_a)))
                producto = col_b

                for mes, col_idx in MESES_COL.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    cantidad = float(val) if val else 0.0
                    registros.append({
                        'cuenta': cuenta_actual,
                        'codigo': codigo,
                        'producto': producto,
                        'negocio': negocio_actual,
                        'categoria': categoria_actual,
                        'anio': 2025,
                        'mes': mes,
                        'mes_num': MESES.index(mes) + 1,  # 1-12
                        'cantidad': cantidad
                    })
            except (ValueError, TypeError):
                pass

    df = pd.DataFrame(registros)
    return df


@st.cache_data(show_spinner=False)
def cargar_datos_2026(archivo_bytes):
    """
    Parsea el Excel jerárquico de 2026 (cuenta -> categoría -> subcategoría -> producto).
    Estructura distinta a 2025: código y nombre van juntos en una sola columna,
    y la cuenta aparece como código numérico de 10 dígitos (no como texto 'cuenta XXXX').
    """
    MESES_COL_2026 = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3,
        'Abril': 5, 'Mayo': 6, 'Junio': 7,
    }
    # mes_num continuo: 2026 sigue después de los 12 meses de 2025 -> Enero 2026 = 13
    MES_NUM_2026 = {
        'Enero': 13, 'Febrero': 14, 'Marzo': 15,
        'Abril': 16, 'Mayo': 17, 'Junio': 18,
    }

    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    registros = []
    cuenta_actual = None
    negocio_actual = None
    categoria_actual = None

    for row in rows:
        col_a = row[0]
        if col_a is None:
            continue
        col_a_str = str(col_a).strip()

        # Detectar cuenta: código puramente numérico de 10 dígitos
        if re.match(r'^\d{10}$', col_a_str):
            cuenta_actual = col_a_str.lstrip('0')
            negocio_actual = None
            categoria_actual = None
            continue

        if not cuenta_actual:
            continue

        # Detectar negocio (2 dígitos) o categoría (4 dígitos)
        if NEGOCIO_PATTERN.match(col_a_str):
            negocio_actual = col_a_str
            categoria_actual = None
            continue
        if CATEGORIA_PATTERN.match(col_a_str):
            categoria_actual = col_a_str
            continue

        # Detectar producto: termina en "(código)"
        m = re.match(r'^(.*)\s+\((\d+)\)$', col_a_str)
        if m:
            producto_limpio = m.group(1).strip()
            codigo = int(m.group(2))

            # Solo procesar columnas que existan en la fila (evita error si el mes aún no tiene datos)
            for mes, col_idx in MESES_COL_2026.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                cantidad = float(val) if val else 0.0
                registros.append({
                    'cuenta': f'cuenta {cuenta_actual}',
                    'codigo': codigo,
                    'producto': col_a_str,
                    'producto_limpio': producto_limpio,
                    'negocio': negocio_actual,
                    'categoria': categoria_actual,
                    'anio': 2026,
                    'mes': mes,
                    'mes_num': MES_NUM_2026[mes],
                    'cantidad': cantidad
                })

    return pd.DataFrame(registros)


@st.cache_data(show_spinner=False)
def cargar_datos_combinado(archivo_bytes):
    """
    Parsea el Excel único que reemplaza a los archivos separados de 2025 y 2026.
    Es una tabla dinámica de Excel con los años como bloques de columnas
    (2025 primero, 2026 a continuación), cada uno con sus meses y columnas
    de 'Total Trimestre X' / 'Total <año>' intercaladas.

    A diferencia de los parsers anteriores (con columnas de mes fijas por
    índice), este detecta el layout de columnas leyendo el encabezado real
    del archivo: ubica la fila 'Etiquetas de fila', mira dos filas arriba
    para saber a qué año pertenece cada bloque de columnas, y solo toma las
    columnas cuyo encabezado es un nombre de mes (así ignora automáticamente
    las columnas de subtotal, sin necesidad de hardcodear índices). Esto lo
    hace robusto a que el archivo crezca mes a mes (p.ej. cuando se sume
    Agosto) o a que se agregue un año más adelante (2027).
    """
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # 1) Ubicar la fila de encabezado de meses ('Etiquetas de fila' en col A)
    header_idx = None
    for i, row in enumerate(rows[:30]):
        if row and row[0] == 'Etiquetas de fila':
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    year_row = rows[header_idx - 2]
    month_row = rows[header_idx]

    # 2) Forward-fill el año a lo largo de las columnas (el Excel solo pone
    #    el año una vez, al principio de cada bloque)
    col_year = {}
    anio_actual = None
    for idx, val in enumerate(year_row):
        if val is not None:
            m = re.match(r'^(20\d{2})$', str(val).strip())
            if m:
                anio_actual = int(m.group(1))
        col_year[idx] = anio_actual

    # 3) Mapear solo las columnas que son un mes real (ignora columnas de Total)
    col_mes = {}
    for idx, val in enumerate(month_row):
        if idx == 0:
            continue
        if val in MESES:
            col_mes[idx] = (val, col_year[idx])

    # 4) Recorrer filas de datos: misma jerarquía cuenta -> negocio -> categoría -> producto
    registros = []
    cuenta_actual = None
    negocio_actual = None
    categoria_actual = None
    ANIO_BASE = 2025  # ancla para el mes_num continuo (Ene25=1, Ene26=13, Ene27=25...)

    for row in rows[header_idx + 1:]:
        col_a = row[0]
        if col_a is None:
            continue
        col_a_str = str(col_a).strip()

        # Detectar cuenta: código puramente numérico de 10 dígitos
        if re.match(r'^\d{10}$', col_a_str):
            cuenta_actual = col_a_str.lstrip('0')
            negocio_actual = None
            categoria_actual = None
            continue

        if not cuenta_actual:
            continue

        if NEGOCIO_PATTERN.match(col_a_str):
            negocio_actual = col_a_str
            categoria_actual = None
            continue
        if CATEGORIA_PATTERN.match(col_a_str):
            categoria_actual = col_a_str
            continue

        # Detectar producto: termina en "(código)"
        m = re.match(r'^(.*)\s+\((\d+)\)$', col_a_str)
        if m:
            producto_limpio = m.group(1).strip()
            codigo = int(m.group(2))

            for col_idx, (mes, anio) in col_mes.items():
                if col_idx >= len(row) or anio is None:
                    continue
                val = row[col_idx]
                cantidad = float(val) if val else 0.0
                mes_num = (anio - ANIO_BASE) * 12 + MESES.index(mes) + 1
                registros.append({
                    'cuenta': f'cuenta {cuenta_actual}',
                    'codigo': codigo,
                    'producto': col_a_str,
                    'producto_limpio': producto_limpio,
                    'negocio': negocio_actual,
                    'categoria': categoria_actual,
                    'anio': anio,
                    'mes': mes,
                    'mes_num': mes_num,
                    'cantidad': cantidad
                })

    return pd.DataFrame(registros)


def limpiar_nombre(nombre):
    """Limpia el nombre del producto removiendo el código al final."""
    import re
    return re.sub(r'\s*\(\d+\)\s*$', '', nombre).strip()


FORMATO_PATTERNS = {
    'Display': re.compile(r'DISPLAY'),
    'Peso (Kg/Grs)': re.compile(r'\bKG\b|X\s*[\d.,]+\s*GRS?\.?\b|X\s*[\d.,]+\s*G\.?\s*$|\d+\s*GS\b|\d+GR\b'),
    'Unidad suelta': re.compile(r'X\s*\d+\s*U(?:NID)?(?:ADES)?\.?\b|\d+\s*U\.?\s*$'),
    'Volumen (Cc/Lt)': re.compile(r'\bCC\b|\bML\b|\bLTS?\b'),
}


def clasificar_formato(nombre):
    """
    Heurística basada en el texto del nombre del producto para inferir su formato
    de venta (display, peso, unidad suelta, volumen). No es un dato garantizado por
    Arcor — es una inferencia de texto, y ~40% de los nombres vienen abreviados por
    el sistema origen y quedan 'Sin clasificar'.
    """
    n = nombre.upper()
    for formato, patron in FORMATO_PATTERNS.items():
        if patron.search(n):
            return formato
    return 'Sin clasificar'


FORMATO_COLOR = {
    'Display': '#818cf8',
    'Peso (Kg/Grs)': '#34d399',
    'Unidad suelta': '#fbbf24',
    'Volumen (Cc/Lt)': '#60a5fa',
    'Sin clasificar': '#6b7280',
}

# Excepciones confirmadas manualmente: cuántas unidades físicas trae cada bulto.
# A diferencia de FORMATO (heurística de texto), esto es un dato exacto confirmado,
# clave para convertir BU -> unidades reales sin sesgo.
PACK_OVERRIDE = {
    15001: 21,   # ALF CHOCO CREAMY 21X
    13757: 21,   # ALF MINITORTA DARK 2 (21X, nombre truncado)
    13357: 21,   # ALF. MINITORTA BCO 2 (21X, nombre truncado)
    13359: 21,   # ALF. MINITORTA BRO 2 (21X, nombre truncado)
    13358: 21,   # ALF. MINITORTA CLA 2 (21X, nombre truncado)
    13360: 21,   # ALF. MINITORTA COC 2 (21X, nombre truncado)
    15489: 21,   # ALF.GOAT NEGRO 21 X 75GR
    6596: 21,    # ALFAJOR COFLER BLOCK 21 x 60 G
    14884: 21,   # BOB TRIPLE 21X60G
    15274: 21,   # BOB TRIPLE BLANCO X60 GRS (21X, confirmado)
}


def badge_pack(codigo, cantidad_bultos):
    """Si el código tiene un pack confirmado, devuelve HTML con la conversión a unidades reales."""
    unidades_x_bulto = PACK_OVERRIDE.get(codigo)
    if not unidades_x_bulto:
        return ''
    total_unidades = cantidad_bultos * unidades_x_bulto
    return f' <span style="color:#f472b6; font-size:0.65rem; font-weight:600;">· ×{unidades_x_bulto}/bulto = {total_unidades:,.0f} uds.</span>'


def etiqueta_mes(mes_num):
    """Convierte mes_num continuo (1-12=2025, 13+=2026) a etiqueta corta tipo 'Ene25'."""
    if mes_num <= 12:
        return f"{MESES[mes_num - 1][:3]}25"
    else:
        return f"{MESES[mes_num - 13][:3]}26"


def proyectar_proximo_pedido(serie_mensual):
    """
    Proyección simple pero robusta:
    - Promedio móvil de los últimos 3 meses con datos
    - Con intervalo ±1 desvío estándar
    """
    valores = [v for v in serie_mensual if v > 0]
    if len(valores) == 0:
        return 0, 0, 0
    recientes = valores[-3:]
    prom = np.mean(recientes)
    std = np.std(recientes) if len(recientes) > 1 else prom * 0.2
    return round(prom, 2), round(max(0, prom - std), 2), round(prom + std, 2)


def calcular_tendencia(serie):
    """Retorna pendiente normalizada de tendencia (-1 a 1)."""
    valores = [v for v in serie if v > 0]
    if len(valores) < 2:
        return 0
    x = np.arange(len(valores))
    coef = np.polyfit(x, valores, 1)
    media = np.mean(valores) if np.mean(valores) != 0 else 1
    return coef[0] / media


def identificar_oportunidades(df_cuenta, df_cadena):
    """
    Identifica productos con potencial de crecimiento en un local:
    - El local compra menos que el promedio del RESTO de la cadena (excluyéndose a sí mismo)
    """
    cuenta_actual = df_cuenta['cuenta'].iloc[0] if not df_cuenta.empty else None

    # Total por producto en el local
    local_total = df_cuenta.groupby(['codigo','producto'])['cantidad'].sum().reset_index()
    local_total.columns = ['codigo','producto','total_local']

    # Cadena SIN el local actual, para que el promedio no se calcule incluyéndose a sí mismo
    df_resto = df_cadena[df_cadena['cuenta'] != cuenta_actual] if cuenta_actual else df_cadena
    n_cuentas_resto = df_resto['cuenta'].nunique()
    if n_cuentas_resto == 0:
        return pd.DataFrame(columns=['codigo','producto','total_local','promedio_cadena','ratio','gap'])

    cadena_total = df_resto.groupby(['codigo','producto'])['cantidad'].sum().reset_index()
    cadena_total['promedio_cadena'] = cadena_total['cantidad'] / n_cuentas_resto
    cadena_total = cadena_total[['codigo','producto','promedio_cadena']]

    merged = local_total.merge(cadena_total, on=['codigo','producto'])
    merged = merged[merged['promedio_cadena'] > 0]
    merged['ratio'] = merged['total_local'] / merged['promedio_cadena']
    merged['gap'] = merged['promedio_cadena'] - merged['total_local']

    # Solo productos con gap positivo (local compra menos que el resto de la cadena)
    oportunidades = merged[merged['gap'] > 0.5].sort_values('gap', ascending=False)
    return oportunidades


# ── Sidebar ────────────────────────────────────────────────────────────────────
APP_VERSION = "v12 · archivo único combinado 2025-2026 · 2026-08-05"

with st.sidebar:
    st.markdown("## 🍬 Golotecas")
    st.markdown("**Análisis de ventas · 2025-2026**")
    st.caption(f"🔖 {APP_VERSION}")
    st.markdown("---")

    archivo = st.file_uploader(
        "Excel ventas mensuales (combinado 2025-2026)",
        type=["xlsx"],
        help="Tabla dinámica única con 2025 y 2026 como bloques de columnas"
    )
    if archivo:
        st.success("✓ Ventas mensuales cargadas")

    st.markdown("---")
    st.markdown("##### Navegación")

    opciones_vista = ["📊 Resumen cadena", "🏪 Análisis por local", "🎯 Próximo pedido", "📈 Comparativa locales", "🗂️ Por negocio", "🗓️ Análisis trimestral"]

    vista = st.radio(
        "Vista",
        opciones_vista,
        label_visibility="collapsed"
    )

# ── Main content ───────────────────────────────────────────────────────────────
if not archivo:
    st.markdown("# 🍬 Golotecas · Análisis de Ventas")
    st.markdown("---")
    col1, col2 = st.columns([2,1])
    with col1:
        st.markdown("""
        ### Cargá el archivo Excel para comenzar

        Esta herramienta analiza las ventas de la cadena de 8 locales y te da:

        - **Proyección del próximo pedido** por local, con intervalo de confianza
        - **Productos a empujar** en cada local según su comportamiento vs. la cadena
        - **Comparativa entre locales** para detectar oportunidades
        - **Tendencias mensuales y trimestrales** por categoría y producto

        Usá el panel izquierdo para subir los archivos Excel.
        """)
    st.stop()

# Cargar datos (archivo único combinado 2025-2026)
with st.spinner("Procesando datos..."):
    df_raw = cargar_datos_combinado(archivo.read())

if df_raw.empty:
    st.error("No se pudieron leer datos del archivo. Verificá el formato.")
    st.stop()

meses_por_anio = df_raw.groupby('anio')['mes'].nunique().to_dict()
resumen_anios = " + ".join(f"{a} ({m} meses)" for a, m in sorted(meses_por_anio.items()))
st.sidebar.caption(f"📅 Cadena: {resumen_anios}")

df_raw['trimestre'] = df_raw['mes'].map(TRIMESTRE_DE_MES)
df_raw['trimestre_label'] = df_raw['anio'].astype(str) + ' T' + df_raw['trimestre'].astype(str)

# Excluir cuenta marginal/ruido (volumen insignificante, no es un local real de la cadena)
df_raw = df_raw[df_raw['cuenta'] != 'cuenta 3080021']

df_raw['formato'] = df_raw['producto_limpio'].apply(clasificar_formato)

CUENTAS = sorted(df_raw['cuenta'].unique())

# ══════════════════════════════════════════════════════════════════════════════
# VISTA 1: RESUMEN CADENA
# ══════════════════════════════════════════════════════════════════════════════
if vista == "📊 Resumen cadena":
    st.markdown("# Resumen · Cadena completa")
    st.markdown("---")

    # Métricas globales
    total_cadena = df_raw['cantidad'].sum()
    total_productos = df_raw['codigo'].nunique()
    total_locales = df_raw['cuenta'].nunique()

    # Mes con más ventas
    por_mes = df_raw.groupby('mes_num')['cantidad'].sum()
    mes_pico_num = por_mes.idxmax()
    mes_pico = etiqueta_mes(mes_pico_num)
    valor_pico = por_mes[mes_pico_num]

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total vendido 2025</div>
            <div class="metric-value">{total_cadena:,.0f}</div>
            <div class="metric-delta delta-neu">BU (unidades base)</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Locales activos</div>
            <div class="metric-value">{total_locales}</div>
            <div class="metric-delta delta-neu">cuentas en la cadena</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Productos distintos</div>
            <div class="metric-value">{total_productos:,}</div>
            <div class="metric-delta delta-neu">SKUs únicos</div>
        </div>""", unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Mes pico</div>
            <div class="metric-value">{mes_pico}</div>
            <div class="metric-delta delta-pos">{valor_pico:,.0f} BU</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown('<div class="section-title">Evolución mensual · Cadena completa</div>', unsafe_allow_html=True)
        evol = df_raw.groupby('mes_num')['cantidad'].sum().reset_index()
        evol['mes'] = evol['mes_num'].apply(etiqueta_mes)
        # Orden cronológico forzado (Streamlit/Altair ordena texto alfabéticamente por defecto)
        evol['mes'] = pd.Categorical(evol['mes'], categories=evol['mes'].tolist(), ordered=True)
        st.bar_chart(evol.set_index('mes')['cantidad'], color="#6366f1", height=240)

    with col_right:
        st.markdown('<div class="section-title">Volumen por local</div>', unsafe_allow_html=True)
        por_local = df_raw.groupby('cuenta')['cantidad'].sum().sort_values(ascending=False)
        por_local.index = [c.replace('cuenta ', '') for c in por_local.index]
        st.bar_chart(por_local, color="#818cf8", height=240)

    # Top productos cadena
    st.markdown('<div class="section-title">Top 15 productos · Cadena</div>', unsafe_allow_html=True)
    top_prods = (df_raw.groupby(['codigo','producto_limpio'])['cantidad']
                 .sum().reset_index()
                 .sort_values('cantidad', ascending=False)
                 .head(15))

    for _, row in top_prods.iterrows():
        pct = row['cantidad'] / total_cadena * 100
        fmt = clasificar_formato(row['producto_limpio'])
        color_fmt = FORMATO_COLOR[fmt]
        pack_html = badge_pack(row['codigo'], row['cantidad'])
        st.markdown(f"""
        <div class="producto-row">
            <span class="producto-nombre">{row['producto_limpio'][:50]} <span style="color:{color_fmt}; font-size:0.65rem; font-weight:600;">· {fmt}</span>{pack_html}</span>
            <span class="producto-valor">{row['cantidad']:,.0f} BU &nbsp;·&nbsp; {pct:.1f}%</span>
        </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# VISTA 2: ANÁLISIS POR LOCAL
# ══════════════════════════════════════════════════════════════════════════════
elif vista == "🏪 Análisis por local":
    st.markdown("# Análisis · Por local")
    st.markdown("---")

    cuenta_sel = st.selectbox(
        "Seleccioná el local",
        CUENTAS,
        format_func=lambda x: x.replace('cuenta ', 'Local ')
    )

    df_local = df_raw[df_raw['cuenta'] == cuenta_sel]

    # Métricas del local
    total_local = df_local['cantidad'].sum()
    total_cadena = df_raw['cantidad'].sum()
    n_locales = df_raw['cuenta'].nunique()
    promedio_cadena = total_cadena / n_locales
    participacion = total_local / total_cadena * 100
    vs_promedio = (total_local - promedio_cadena) / promedio_cadena * 100

    col1, col2, col3 = st.columns(3)
    with col1:
        signo = "+" if vs_promedio >= 0 else ""
        clase = "delta-pos" if vs_promedio >= 0 else "delta-neg"
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total local 2025</div>
            <div class="metric-value">{total_local:,.0f}</div>
            <div class="metric-delta {clase}">{signo}{vs_promedio:.1f}% vs promedio cadena</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Participación en cadena</div>
            <div class="metric-value">{participacion:.1f}%</div>
            <div class="metric-delta delta-neu">del total cadena</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        prods_activos = (df_local.groupby('codigo')['cantidad'].sum() > 0).sum()
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Productos activos</div>
            <div class="metric-value">{prods_activos}</div>
            <div class="metric-delta delta-neu">con al menos 1 pedido</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.markdown(f'<div class="section-title">Evolución mensual · {cuenta_sel.replace("cuenta ", "Local ")}</div>', unsafe_allow_html=True)
        evol_local = df_local.groupby('mes_num')['cantidad'].sum().reset_index()
        evol_local['mes'] = evol_local['mes_num'].apply(etiqueta_mes)
        evol_local['mes'] = pd.Categorical(evol_local['mes'], categories=evol_local['mes'].tolist(), ordered=True)
        st.bar_chart(evol_local.set_index('mes')['cantidad'], color="#6366f1", height=220)

    with col_right:
        st.markdown('<div class="section-title">Top 5 categorías</div>', unsafe_allow_html=True)
        # Extraer categoría del nombre de producto usando el codigo (aproximado)
        top5 = (df_local.groupby('producto_limpio')['cantidad']
                .sum().sort_values(ascending=False).head(5))
        for prod, val in top5.items():
            pct = val / total_local * 100
            st.markdown(f"""
            <div class="producto-row">
                <span class="producto-nombre">{prod[:40]}</span>
                <span class="producto-valor">{val:.0f} · {pct:.0f}%</span>
            </div>""", unsafe_allow_html=True)

    # Oportunidades
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">Productos con oportunidad · Este local compra menos que el promedio de la cadena</div>', unsafe_allow_html=True)

    oport = identificar_oportunidades(df_local, df_raw)
    if not oport.empty:
        oport['producto_limpio'] = oport['producto'].apply(limpiar_nombre)
        for _, row in oport.head(10).iterrows():
            gap = row['gap']
            ratio = row['ratio']
            if ratio < 0.3:
                tag = '<span class="tag-rojo">Muy por debajo</span>'
            elif ratio < 0.7:
                tag = '<span class="tag-amarillo">Por debajo</span>'
            else:
                tag = '<span class="tag-verde">Leve oportunidad</span>'

            st.markdown(f"""
            <div class="producto-row">
                <span class="producto-nombre">{row['producto_limpio'][:50]} &nbsp; {tag}</span>
                <span class="producto-valor">gap {gap:.1f} BU</span>
            </div>""", unsafe_allow_html=True)
    else:
        st.info("Este local está por encima del promedio en todos los productos activos.")


# ══════════════════════════════════════════════════════════════════════════════
# VISTA 3: PRÓXIMO PEDIDO
# ══════════════════════════════════════════════════════════════════════════════
elif vista == "🎯 Próximo pedido":
    st.markdown("# Proyección · Próximo pedido")
    st.markdown("---")

    cuenta_sel = st.selectbox(
        "Seleccioná el local",
        CUENTAS,
        format_func=lambda x: x.replace('cuenta ', 'Local ')
    )

    df_local = df_raw[df_raw['cuenta'] == cuenta_sel]

    # Proyección total del local
    meses_disp = sorted(df_raw['mes_num'].unique())
    serie_total = [df_local[df_local['mes_num'] == m]['cantidad'].sum() for m in meses_disp]
    proy_total, low, high = proyectar_proximo_pedido(serie_total)

    st.markdown(f"""
    <div class="proyeccion-card">
        <div class="proyeccion-title">Proyección próximo pedido · {cuenta_sel.replace("cuenta ", "Local ")}</div>
        <div class="proyeccion-valor">{proy_total:,.0f} BU</div>
        <div class="proyeccion-rango">Rango estimado: {low:,.0f} – {high:,.0f} BU &nbsp;·&nbsp; Basado en promedio móvil 3 meses</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Proyección por producto (los más relevantes)
    st.markdown('<div class="section-title">Proyección por producto · Top 20 por volumen</div>', unsafe_allow_html=True)

    top_prods = (df_local.groupby(['codigo','producto_limpio'])['cantidad']
                 .sum().reset_index()
                 .sort_values('cantidad', ascending=False)
                 .head(20))

    col1, col2 = st.columns(2)
    for idx, (_, prod_row) in enumerate(top_prods.iterrows()):
        df_prod = df_local[df_local['codigo'] == prod_row['codigo']]
        serie = [df_prod[df_prod['mes_num'] == m]['cantidad'].sum() for m in meses_disp]
        proy, lo, hi = proyectar_proximo_pedido(serie)

        # Calcular tendencia
        tend = calcular_tendencia(serie)
        if tend > 0.1:
            tend_txt = '<span class="tag-verde">↑ sube</span>'
        elif tend < -0.1:
            tend_txt = '<span class="tag-rojo">↓ baja</span>'
        else:
            tend_txt = '<span class="tag-amarillo">→ estable</span>'

        contenido = f"""
        <div class="producto-row" style="border-left-color: {'#34d399' if tend > 0.1 else '#f87171' if tend < -0.1 else '#fbbf24'}">
            <span class="producto-nombre">{prod_row['producto_limpio'][:45]} &nbsp; {tend_txt}</span>
            <span class="producto-valor">{proy:.1f} BU</span>
        </div>"""

        if idx % 2 == 0:
            with col1:
                st.markdown(contenido, unsafe_allow_html=True)
        else:
            with col2:
                st.markdown(contenido, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:0.72rem; color:#4b5563; padding: 0.8rem; background:#111827; border-radius:8px;">
    📌 La proyección usa promedio móvil de los últimos 3 meses con datos. El rango refleja la variabilidad histórica reciente. 
    Para pedidos irregulares, tomá la proyección como referencia base, no como número exacto.
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# VISTA 4: COMPARATIVA LOCALES
# ══════════════════════════════════════════════════════════════════════════════
elif vista == "📈 Comparativa locales":
    st.markdown("# Comparativa · Todos los locales")
    st.markdown("---")

    # Tabla resumen por local
    resumen = []
    for cuenta in CUENTAS:
        df_c = df_raw[df_raw['cuenta'] == cuenta]
        total = df_c['cantidad'].sum()
        prods_activos = (df_c.groupby('codigo')['cantidad'].sum() > 0).sum()
        meses_disp_c = sorted(df_raw['mes_num'].unique())
        serie = [df_c[df_c['mes_num'] == m]['cantidad'].sum() for m in meses_disp_c]
        proy, _, _ = proyectar_proximo_pedido(serie)
        tend = calcular_tendencia(serie)
        top_prod = (df_c.groupby('producto_limpio')['cantidad'].sum()
                    .sort_values(ascending=False).index[0] if not df_c.empty else '-')
        resumen.append({
            'Local': cuenta.replace('cuenta ', 'Local '),
            'Total 2025': total,
            'Productos activos': prods_activos,
            'Próx. pedido (est.)': proy,
            'Tendencia': tend,
            'Producto #1': top_prod[:40]
        })

    df_resumen = pd.DataFrame(resumen).sort_values('Total 2025', ascending=False)

    st.markdown('<div class="section-title">Ranking de locales por volumen</div>', unsafe_allow_html=True)
    for _, row in df_resumen.iterrows():
        tend = row['Tendencia']
        if tend > 0.1:
            tend_txt = '<span class="tag-verde">↑ tendencia positiva</span>'
        elif tend < -0.1:
            tend_txt = '<span class="tag-rojo">↓ tendencia negativa</span>'
        else:
            tend_txt = '<span class="tag-amarillo">→ estable</span>'

        st.markdown(f"""
        <div class="metric-card" style="margin-bottom:0.6rem;">
            <div style="display:flex; justify-content:space-between; align-items:center;">
                <div>
                    <div style="font-size:1rem; font-weight:700; color:#f9fafb;">{row['Local']}</div>
                    <div style="font-size:0.72rem; color:#6b7280; margin-top:0.2rem;">
                        Top producto: {row['Producto #1']} &nbsp; {tend_txt}
                    </div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:1.4rem; font-weight:700; color:#a5b4fc;">{row['Total 2025']:,.0f} BU</div>
                    <div style="font-size:0.72rem; color:#6b7280;">
                        {row['Productos activos']} SKUs · próx. pedido ~{row['Próx. pedido (est.)']:,.0f}
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Evolución mensual superpuesta
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">Evolución mensual · Todos los locales</div>', unsafe_allow_html=True)

    evol_pivot = df_raw.groupby(['cuenta','mes_num'])['cantidad'].sum().reset_index()
    evol_pivot['cuenta'] = evol_pivot['cuenta'].str.replace('cuenta ', 'L')
    pivot = evol_pivot.pivot(index='mes_num', columns='cuenta', values='cantidad').fillna(0)
    etiquetas_orden = [etiqueta_mes(i) for i in pivot.index]
    pivot.index = pd.CategoricalIndex(etiquetas_orden, categories=etiquetas_orden, ordered=True)
    st.line_chart(pivot, height=300)

    # Producto estrella por local
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">Producto estrella por local</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    for idx, cuenta in enumerate(CUENTAS):
        df_c = df_raw[df_raw['cuenta'] == cuenta]
        top3 = (df_c.groupby('producto_limpio')['cantidad']
                .sum().sort_values(ascending=False).head(3))
        contenido = f"""
        <div class="metric-card" style="margin-bottom:0.4rem;">
            <div class="metric-label">{cuenta.replace('cuenta ', 'Local ')}</div>
            {''.join([f'<div style="font-size:0.78rem; color:#d1d5db; padding:0.1rem 0;">{p[:45]} <span style="color:#6366f1;">{v:.0f} BU</span></div>' for p, v in top3.items()])}
        </div>"""
        if idx % 2 == 0:
            with col1:
                st.markdown(contenido, unsafe_allow_html=True)
        else:
            with col2:
                st.markdown(contenido, unsafe_allow_html=True)



# ══════════════════════════════════════════════════════════════════════════════
# VISTA 5: POR NEGOCIO
# ══════════════════════════════════════════════════════════════════════════════
elif vista == "🗂️ Por negocio":
    st.markdown("# Movimientos · Por negocio")
    st.markdown("---")

    if 'negocio' not in df_raw.columns or df_raw['negocio'].isna().all():
        st.warning("El archivo cargado no tiene información de negocio/categoría detectada.")
        st.stop()

    # Filtro opcional por local y por formato
    col_local, col_formato = st.columns([1, 1])
    with col_local:
        filtro_local = st.selectbox(
            "Local",
            ["Cadena completa"] + CUENTAS,
            format_func=lambda x: x if x == "Cadena completa" else x.replace('cuenta ', 'Local ')
        )
    with col_formato:
        filtro_formato = st.selectbox(
            "Formato de venta",
            ["Todos"] + list(FORMATO_COLOR.keys()),
            help="Inferido del nombre del producto (heurística de texto, no un dato garantizado por Arcor)"
        )

    df_neg = df_raw if filtro_local == "Cadena completa" else df_raw[df_raw['cuenta'] == filtro_local]
    df_neg = df_neg[df_neg['negocio'].notna()]
    if filtro_formato != "Todos":
        df_neg = df_neg[df_neg['formato'] == filtro_formato]

    # Resumen por negocio
    resumen_negocio = (df_neg.groupby('negocio')['cantidad']
                       .sum().sort_values(ascending=False).reset_index())
    total_general = resumen_negocio['cantidad'].sum()

    st.markdown('<div class="section-title">Resumen · Todos los negocios</div>', unsafe_allow_html=True)
    for _, row_n in resumen_negocio.iterrows():
        pct = row_n['cantidad'] / total_general * 100 if total_general else 0
        st.markdown(f"""
        <div class="metric-card" style="margin-bottom:0.5rem;">
            <div style="display:flex; justify-content:space-between; align-items:center;">
                <div style="font-size:0.95rem; font-weight:700; color:#f9fafb;">{row_n['negocio']}</div>
                <div style="text-align:right;">
                    <div style="font-size:1.3rem; font-weight:700; color:#a5b4fc;">{row_n['cantidad']:,.0f} BU</div>
                    <div style="font-size:0.72rem; color:#6b7280;">{pct:.1f}% del total</div>
                </div>
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Drill-down: elegir negocio y ver sus categorías + productos
    st.markdown('<div class="section-title">Detalle · Categorías y productos dentro de un negocio</div>', unsafe_allow_html=True)

    negocios_disp = sorted(df_neg['negocio'].unique())
    negocio_sel = st.selectbox("Elegí un negocio para ver el detalle", negocios_disp)

    df_negocio_sel = df_neg[df_neg['negocio'] == negocio_sel]

    # Categorías dentro del negocio elegido
    cat_resumen = (df_negocio_sel[df_negocio_sel['categoria'].notna()]
                   .groupby('categoria')['cantidad'].sum()
                   .sort_values(ascending=False).reset_index())

    total_negocio = df_negocio_sel['cantidad'].sum()

    col_cat, col_prod = st.columns(2)

    with col_cat:
        st.markdown(f'<div style="font-size:0.72rem; font-weight:700; letter-spacing:0.08em; color:#818cf8; text-transform:uppercase; margin-bottom:0.6rem;">Categorías en {negocio_sel}</div>', unsafe_allow_html=True)
        for _, row_c in cat_resumen.iterrows():
            pct_c = row_c['cantidad'] / total_negocio * 100 if total_negocio else 0
            st.markdown(f"""
            <div class="producto-row" style="border-left-color:#818cf8;">
                <span class="producto-nombre">{row_c['categoria']}</span>
                <span class="producto-valor">{row_c['cantidad']:,.0f} BU · {pct_c:.0f}%</span>
            </div>""", unsafe_allow_html=True)

    with col_prod:
        st.markdown(f'<div style="font-size:0.72rem; font-weight:700; letter-spacing:0.08em; color:#34d399; text-transform:uppercase; margin-bottom:0.6rem;">Top 15 productos en {negocio_sel}</div>', unsafe_allow_html=True)
        top_prod_negocio = (df_negocio_sel.groupby(['codigo', 'producto_limpio'])['cantidad']
                            .sum().reset_index().sort_values('cantidad', ascending=False).head(15))
        for _, r in top_prod_negocio.iterrows():
            prod, val, cod = r['producto_limpio'], r['cantidad'], r['codigo']
            pct_p = val / total_negocio * 100 if total_negocio else 0
            fmt = clasificar_formato(prod)
            color_fmt = FORMATO_COLOR[fmt]
            pack_html = badge_pack(cod, val)
            st.markdown(f"""
            <div class="producto-row" style="border-left-color:#34d399;">
                <span class="producto-nombre">{prod[:32]} <span style="color:{color_fmt}; font-size:0.62rem; font-weight:600;">· {fmt}</span>{pack_html}</span>
                <span class="producto-valor">{val:,.0f} · {pct_p:.0f}%</span>
            </div>""", unsafe_allow_html=True)

    # Filtro adicional: elegir una categoría específica y ver solo sus productos
    if not cat_resumen.empty:
        st.markdown("<br>", unsafe_allow_html=True)
        categoria_sel = st.selectbox(
            "O elegí una categoría puntual para ver todos sus productos",
            ["(ninguna)"] + list(cat_resumen['categoria'])
        )
        if categoria_sel != "(ninguna)":
            df_cat_sel = df_negocio_sel[df_negocio_sel['categoria'] == categoria_sel]
            prods_cat = (df_cat_sel.groupby(['codigo', 'producto_limpio'])['cantidad']
                        .sum().reset_index().sort_values('cantidad', ascending=False))
            st.markdown(f'<div class="section-title">Todos los productos en {categoria_sel}</div>', unsafe_allow_html=True)
            for _, r in prods_cat.iterrows():
                prod, val, cod = r['producto_limpio'], r['cantidad'], r['codigo']
                pack_html = badge_pack(cod, val)
                st.markdown(f"""
                <div class="producto-row">
                    <span class="producto-nombre">{prod[:55]}{pack_html}</span>
                    <span class="producto-valor">{val:,.0f} BU</span>
                </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# VISTA 6: ANÁLISIS TRIMESTRAL
# ══════════════════════════════════════════════════════════════════════════════
elif vista == "🗓️ Análisis trimestral":
    st.markdown("# Análisis · Trimestral y estacionalidad")
    st.markdown("---")

    col_local, _ = st.columns([1, 2])
    with col_local:
        filtro_local_tri = st.selectbox(
            "Local",
            ["Cadena completa"] + CUENTAS,
            format_func=lambda x: x if x == "Cadena completa" else x.replace('cuenta ', 'Local '),
            key="sel_trimestral"
        )

    df_tri = df_raw if filtro_local_tri == "Cadena completa" else df_raw[df_raw['cuenta'] == filtro_local_tri]

    # ── Evolución por trimestre (todos los trimestres disponibles en orden) ──
    st.markdown('<div class="section-title">Volumen por trimestre</div>', unsafe_allow_html=True)

    orden_trimestres = (df_tri[['anio', 'trimestre', 'trimestre_label']]
                        .drop_duplicates()
                        .sort_values(['anio', 'trimestre']))
    por_trimestre = df_tri.groupby('trimestre_label')['cantidad'].sum()
    por_trimestre = por_trimestre.reindex(orden_trimestres['trimestre_label'])

    st.bar_chart(por_trimestre, color="#6366f1", height=260)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Volumen mensual: para descartar meses puntuales de grandes eventos ──
    st.markdown('<div class="section-title">Volumen mensual · detectar meses con eventos puntuales</div>', unsafe_allow_html=True)
    st.caption("Cada mes se compara contra el promedio de los otros 2 meses de su mismo trimestre — así se distingue una compra grande y puntual de una tendencia real del trimestre.")

    orden_meses = (df_tri[['anio', 'mes_num', 'mes', 'trimestre_label']]
                  .drop_duplicates()
                  .sort_values('mes_num'))
    por_mes_total = df_tri.groupby('mes_num')['cantidad'].sum()

    serie_mensual = por_mes_total.reindex(orden_meses['mes_num'])
    serie_mensual.index = orden_meses.apply(lambda r: f"{r['mes'][:3]} {r['anio']}", axis=1).values

    st.bar_chart(serie_mensual, color="#a78bfa", height=240)

    for _, fila_mes in orden_meses.iterrows():
        mn = fila_mes['mes_num']
        tl = fila_mes['trimestre_label']
        vol_mes = por_mes_total.get(mn, 0.0)

        # Comparar contra el promedio de los otros meses del mismo trimestre
        meses_mismo_trim = orden_meses[orden_meses['trimestre_label'] == tl]['mes_num'].tolist()
        otros_meses = [m for m in meses_mismo_trim if m != mn]
        vols_otros = [por_mes_total.get(m, 0.0) for m in otros_meses]
        prom_otros = sum(vols_otros) / len(vols_otros) if vols_otros else None

        alerta = ''
        es_pico = False
        if prom_otros and prom_otros > 0:
            desvio_pct = (vol_mes - prom_otros) / prom_otros * 100
            if desvio_pct >= 50:
                es_pico = True
                alerta = f' <span class="tag-amarillo">⚠ {desvio_pct:.0f}% arriba del promedio del trimestre</span>'

        etiqueta = f"{fila_mes['mes']} {fila_mes['anio']}"
        st.markdown(f"""
        <div class="producto-row">
            <span class="producto-nombre">{etiqueta} <span style="color:#6b7280; font-size:0.7rem;">({tl})</span></span>
            <span class="producto-valor">{vol_mes:,.0f} BU{alerta}</span>
        </div>""", unsafe_allow_html=True)

        if es_pico:
            if filtro_local_tri == "Cadena completa":
                top_contrib = (df_tri[df_tri['mes_num'] == mn]
                              .groupby('cuenta')['cantidad'].sum()
                              .sort_values(ascending=False).head(3))
                detalle_txt = " · ".join(f"{n.replace('cuenta ', 'Local ')} ({v:,.0f} BU)" for n, v in top_contrib.items())
                st.caption(f"↳ Locales que más aportaron ese mes: {detalle_txt}")
            else:
                top_contrib = (df_tri[df_tri['mes_num'] == mn]
                              .groupby('producto_limpio')['cantidad'].sum()
                              .sort_values(ascending=False).head(3))
                detalle_txt = " · ".join(f"{n[:30]} ({v:,.0f} BU)" for n, v in top_contrib.items())
                st.caption(f"↳ Productos que más aportaron ese mes: {detalle_txt}")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Comparativa año contra año, mismo trimestre ──
    st.markdown('<div class="section-title">Comparativa · Mismo trimestre, año contra año</div>', unsafe_allow_html=True)

    trimestres_presentes = sorted(df_tri['trimestre'].unique())
    anios_presentes = sorted(df_tri['anio'].unique())

    if len(anios_presentes) >= 2:
        for t in trimestres_presentes:
            df_t = df_tri[df_tri['trimestre'] == t]
            por_anio_t = df_t.groupby('anio')['cantidad'].sum()

            # Solo comparar si el trimestre tiene datos en más de un año
            if len(por_anio_t) < 2:
                continue

            anio_base = min(por_anio_t.index)
            anio_comp = max(por_anio_t.index)
            val_base = por_anio_t[anio_base]
            val_comp = por_anio_t[anio_comp]
            var_pct = (val_comp - val_base) / val_base * 100 if val_base else 0

            clase = "delta-pos" if var_pct >= 0 else "delta-neg"
            signo = "+" if var_pct >= 0 else ""
            color_barra = "#34d399" if var_pct >= 0 else "#f87171"

            st.markdown(f"""
            <div class="metric-card" style="margin-bottom:0.5rem;">
                <div style="display:flex; justify-content:space-between; align-items:center;">
                    <div>
                        <div style="font-size:0.95rem; font-weight:700; color:#f9fafb;">Trimestre {t}</div>
                        <div style="font-size:0.72rem; color:#6b7280; margin-top:0.15rem;">
                            {anio_base} T{t}: {val_base:,.0f} BU &nbsp;→&nbsp; {anio_comp} T{t}: {val_comp:,.0f} BU
                        </div>
                    </div>
                    <div style="text-align:right;">
                        <div style="font-size:1.4rem; font-weight:700; color:{color_barra};">{signo}{var_pct:.1f}%</div>
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)
    else:
        st.info("Con un solo año cargado no hay comparativa año contra año todavía. Sumá el archivo de 2026 (o años siguientes) para habilitar esta vista.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Estacionalidad por producto: elegir un trimestre y ver qué domina ──
    st.markdown('<div class="section-title">Productos que dominan cada trimestre</div>', unsafe_allow_html=True)

    trimestre_sel = st.selectbox(
        "Elegí un trimestre para ver su top de productos",
        sorted(orden_trimestres['trimestre_label'].unique(), key=lambda x: (x.split(' T')[0], x.split(' T')[1]))
    )

    df_trim_sel = df_tri[df_tri['trimestre_label'] == trimestre_sel]
    top_prod_trim = (df_trim_sel.groupby('producto_limpio')['cantidad']
                     .sum().sort_values(ascending=False).head(25))
    total_trim_sel = df_trim_sel['cantidad'].sum()

    for prod, val in top_prod_trim.items():
        pct = val / total_trim_sel * 100 if total_trim_sel else 0
        st.markdown(f"""
        <div class="producto-row">
            <span class="producto-nombre">{prod[:45]}</span>
            <span class="producto-valor">{val:,.0f} BU · {pct:.0f}%</span>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Evolución de un producto puntual a través de todos los trimestres ──
    st.markdown('<div class="section-title">Evolución de un producto puntual</div>', unsafe_allow_html=True)
    st.caption("Elegí un producto y mirá su volumen y su puesto en el ranking en cada trimestre, incluso cuando no entra en el top 25 — así se detectan desvíos puntuales.")

    productos_disponibles = sorted(df_tri['producto_limpio'].unique())
    producto_foco = st.selectbox(
        "Producto a analizar",
        productos_disponibles,
        key="sel_producto_foco"
    )

    orden_trim_list = list(orden_trimestres['trimestre_label'])
    filas_foco = []
    for tl in orden_trim_list:
        df_t_all = df_tri[df_tri['trimestre_label'] == tl]
        ranking_t = df_t_all.groupby('producto_limpio')['cantidad'].sum().sort_values(ascending=False)
        n_productos_t = len(ranking_t)
        if producto_foco in ranking_t.index:
            vol = ranking_t[producto_foco]
            puesto = ranking_t.index.get_loc(producto_foco) + 1
        else:
            vol = 0.0
            puesto = None
        filas_foco.append({'trimestre': tl, 'volumen': vol, 'puesto': puesto, 'n_productos': n_productos_t})

    df_foco = pd.DataFrame(filas_foco)

    st.bar_chart(df_foco.set_index('trimestre')['volumen'], color="#818cf8", height=220)

    for i, row_f in df_foco.iterrows():
        if row_f['puesto'] is None:
            puesto_txt = "sin ventas este trimestre"
            color_puesto = "#f87171"
        elif row_f['puesto'] <= 25:
            puesto_txt = f"puesto #{row_f['puesto']} de {row_f['n_productos']}"
            color_puesto = "#34d399"
        else:
            puesto_txt = f"puesto #{row_f['puesto']} de {row_f['n_productos']} · fuera del top 25"
            color_puesto = "#fbbf24"

        # Marcar caída fuerte vs el trimestre inmediatamente anterior (posible desvío puntual)
        alerta_txt = ''
        if i > 0:
            vol_prev = df_foco.iloc[i - 1]['volumen']
            if vol_prev > 0:
                var_pct_f = (row_f['volumen'] - vol_prev) / vol_prev * 100
                if var_pct_f <= -30:
                    alerta_txt = f' <span class="tag-rojo">⚠ {var_pct_f:.0f}% vs trim. anterior</span>'

        st.markdown(f"""
        <div class="producto-row">
            <span class="producto-nombre">{row_f['trimestre']}</span>
            <span class="producto-valor" style="color:{color_puesto};">{row_f['volumen']:,.0f} BU · {puesto_txt}{alerta_txt}</span>
        </div>""", unsafe_allow_html=True)

    st.caption("💡 El ⚠ marca una caída de 30% o más contra el trimestre inmediatamente anterior — no explica la causa, es una señal para revisar con tu criterio si hubo algo puntual (quiebre de stock, cambio de precio, estacionalidad, etc.).")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Desglose mensual: mismo producto elegido, abierto mes a mes dentro del trimestre elegido arriba ──
    st.markdown(f'<div class="section-title">Desglose mensual · {producto_foco[:40]} en {trimestre_sel}</div>', unsafe_allow_html=True)

    df_prod_mes = df_tri[
        (df_tri['trimestre_label'] == trimestre_sel) &
        (df_tri['producto_limpio'] == producto_foco)
    ]

    if df_prod_mes.empty:
        st.caption(f"Sin datos de \"{producto_foco}\" en {trimestre_sel}.")
    else:
        por_mes_prod = (df_prod_mes.groupby('mes_num')
                        .agg(mes=('mes', 'first'), cantidad=('cantidad', 'sum'))
                        .sort_index())
        codigo_prod = df_prod_mes['codigo'].iloc[0]
        total_prod_trim = por_mes_prod['cantidad'].sum()

        st.bar_chart(por_mes_prod.set_index('mes')['cantidad'], color="#34d399", height=200)

        for _, row_m in por_mes_prod.iterrows():
            pct_m = row_m['cantidad'] / total_prod_trim * 100 if total_prod_trim else 0
            pack_html = badge_pack(codigo_prod, row_m['cantidad'])
            st.markdown(f"""
            <div class="producto-row">
                <span class="producto-nombre">{row_m['mes']}</span>
                <span class="producto-valor">{row_m['cantidad']:,.2f} BU · {pct_m:.0f}% del trimestre{pack_html}</span>
            </div>""", unsafe_allow_html=True)

    # ── Comparar el mix de productos de este trimestre vs mismo trimestre año anterior ──
    trimestre_num_sel = int(trimestre_sel.split(' T')[1])
    anio_num_sel = int(trimestre_sel.split(' T')[0])
    anio_anterior = anio_num_sel - 1

    df_trim_anterior = df_tri[(df_tri['trimestre'] == trimestre_num_sel) & (df_tri['anio'] == anio_anterior)]

    if not df_trim_anterior.empty:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="section-title">Qué cambió · T{trimestre_num_sel} {anio_anterior} → T{trimestre_num_sel} {anio_num_sel}</div>', unsafe_allow_html=True)

        actual_por_prod = df_trim_sel.groupby('producto_limpio')['cantidad'].sum()
        anterior_por_prod = df_trim_anterior.groupby('producto_limpio')['cantidad'].sum()

        comparativa = pd.DataFrame({'actual': actual_por_prod, 'anterior': anterior_por_prod}).fillna(0)
        comparativa['var_abs'] = comparativa['actual'] - comparativa['anterior']
        comparativa = comparativa.sort_values('var_abs', ascending=False)

        col_sube_t, col_baja_t = st.columns(2)
        with col_sube_t:
            st.markdown('<div style="font-size:0.72rem; font-weight:700; letter-spacing:0.08em; color:#34d399; text-transform:uppercase; margin-bottom:0.6rem;">↑ Crecieron más</div>', unsafe_allow_html=True)
            for prod, row_c in comparativa[comparativa['var_abs'] > 0].head(10).iterrows():
                es_nuevo = row_c['anterior'] == 0
                tag_nuevo = ' <span class="tag-amarillo">nuevo</span>' if es_nuevo else ''
                st.markdown(f"""
                <div class="producto-row" style="border-left-color:#34d399;">
                    <span class="producto-nombre">{prod[:34]}{tag_nuevo}</span>
                    <span class="producto-valor" style="color:#34d399;">+{row_c['var_abs']:.0f} BU</span>
                </div>""", unsafe_allow_html=True)
        with col_baja_t:
            st.markdown('<div style="font-size:0.72rem; font-weight:700; letter-spacing:0.08em; color:#f87171; text-transform:uppercase; margin-bottom:0.6rem;">↓ Bajaron más</div>', unsafe_allow_html=True)
            for prod, row_c in comparativa[comparativa['var_abs'] < 0].sort_values('var_abs').head(10).iterrows():
                es_discontinuado = row_c['actual'] == 0
                tag_disc = ' <span class="tag-rojo">discontinuado</span>' if es_discontinuado else ''
                st.markdown(f"""
                <div class="producto-row" style="border-left-color:#f87171;">
                    <span class="producto-nombre">{prod[:34]}{tag_disc}</span>
                    <span class="producto-valor" style="color:#f87171;">{row_c['var_abs']:.0f} BU</span>
                </div>""", unsafe_allow_html=True)
        st.caption("💡 \"nuevo\" = no existía en el trimestre anterior (no es crecimiento orgánico). \"discontinuado\" = no tuvo ventas en el trimestre actual.")
    else:
        st.caption(f"No hay datos de T{trimestre_num_sel} {anio_anterior} para comparar contra este trimestre.")
